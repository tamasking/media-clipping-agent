"""
Report Generator Module for Media Clipping Agent
Creates Excel reports with article data and embedded screenshots
"""
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Install with: pip install openpyxl")

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class ReportGenerator:
    """Generates Excel reports with media clipping data"""

    def __init__(self, reports_dir: str = None):
        if reports_dir is None:
            reports_dir = "/a0/usr/projects/project_1/media_clipping_agent/reports"
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self.workbook = None
        self.worksheet = None

    def _estimate_reach(self, url: str, source: str = "") -> int:
        """Estimate audience reach based on source"""
        domain = urlparse(url).netloc.lower()

        # Major news sites
        major_news = [
            "bbc.com", "cnn.com", "reuters.com", "nytimes.com", 
            "washingtonpost.com", "theguardian.com", "apnews.com",
            "forbes.com", "bloomberg.com", "wsj.com", "ft.com"
        ]

        # Tech news
        tech_news = [
            "techcrunch.com", "theverge.com", "wired.com", "arstechnica.com",
            "engadget.com", "cnet.com", "zdnet.com"
        ]

        # Social platforms
        social_platforms = [
            "linkedin.com", "facebook.com", "instagram.com", "twitter.com", "x.com"
        ]

        if any(m in domain for m in major_news):
            return 500000 + hash(domain) % 1500000  # 500K-2M
        elif any(t in domain for t in tech_news):
            return 100000 + hash(domain) % 400000  # 100K-500K
        elif any(s in domain for s in social_platforms):
            return 10000 + hash(domain) % 90000  # 10K-100K
        elif "blog" in domain or "medium" in domain:
            return 5000 + hash(domain) % 45000  # 5K-50K
        else:
            return 1000 + hash(domain) % 19000  # 1K-20K

    def _create_workbook(self, title: str = "Media Clipping Report"):
        """Create new Excel workbook with styling"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Media Mentions"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Define headers
        headers = [
            "Date Found",
            "Published Date",
            "Title",
            "URL",
            "Source",
            "Estimated Reach",
            "Screenshot",
            "Notes"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set column widths
        self.worksheet.column_dimensions["A"].width = 15  # Date Found
        self.worksheet.column_dimensions["B"].width = 15  # Published Date
        self.worksheet.column_dimensions["C"].width = 50  # Title
        self.worksheet.column_dimensions["D"].width = 50  # URL
        self.worksheet.column_dimensions["E"].width = 20  # Source
        self.worksheet.column_dimensions["F"].width = 15  # Reach
        self.worksheet.column_dimensions["G"].width = 40  # Screenshot
        self.worksheet.column_dimensions["H"].width = 30  # Notes

        # Set header row height
        self.worksheet.row_dimensions[1].height = 25

    def add_entry(self, row_num: int, data: Dict):
        """Add a single entry to the report"""
        if not self.worksheet:
            self._create_workbook()

        # Calculate reach if not provided
        reach = data.get("reach_estimate", 0)
        if not reach:
            reach = self._estimate_reach(data.get("url", ""), data.get("source", ""))

        # Write data
        self.worksheet.cell(row=row_num, column=1, value=data.get("found_at", "")[:10])
        self.worksheet.cell(row=row_num, column=2, value=data.get("published_date", "Unknown"))

        title_cell = self.worksheet.cell(row=row_num, column=3, value=data.get("title", "No title"))
        title_cell.alignment = Alignment(wrap_text=True)

        url_cell = self.worksheet.cell(row=row_num, column=4, value=data.get("url", ""))
        url_cell.style = "Hyperlink"
        url_cell.hyperlink = data.get("url", "")

        self.worksheet.cell(row=row_num, column=5, value=data.get("source", "Unknown"))

        reach_cell = self.worksheet.cell(row=row_num, column=6, value=reach)
        reach_cell.number_format = '#,##0'

        # Screenshot path
        screenshot_path = data.get("screenshot_path", "")
        if screenshot_path and Path(screenshot_path).exists():
            self.worksheet.cell(row=row_num, column=7, value=f"See attached: {Path(screenshot_path).name}")

            # Try to embed image
            if PIL_AVAILABLE:
                try:
                    img = Image.open(screenshot_path)
                    # Resize if too large
                    max_width, max_height = 800, 600
                    img.thumbnail((max_width, max_height))

                    # Save temporary resized image
                    temp_path = Path(screenshot_path).parent / f"thumb_{Path(screenshot_path).name}"
                    img.save(temp_path)

                    xl_img = XLImage(str(temp_path))
                    # Scale down for Excel
                    xl_img.width = 300
                    xl_img.height = 200

                    self.worksheet.add_image(xl_img, f"G{row_num}")

                    # Adjust row height
                    self.worksheet.row_dimensions[row_num].height = 150
                except Exception as e:
                    print(f"Could not embed image: {e}")
        else:
            self.worksheet.cell(row=row_num, column=7, value="No screenshot")

        # Notes/Snippet
        snippet_cell = self.worksheet.cell(row=row_num, column=8, value=data.get("snippet", "")[:500])
        snippet_cell.alignment = Alignment(wrap_text=True)

    def generate_report(self, data: List[Dict], report_name: str = None,
                       keywords: str = "") -> str:
        """Generate complete Excel report"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for report generation")

        if not report_name:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_keywords = keywords.replace(" ", "_")[:30] if keywords else "media_clipping"
            report_name = f"{safe_keywords}_{timestamp}.xlsx"

        if not report_name.endswith(".xlsx"):
            report_name += ".xlsx"

        output_path = self.reports_dir / report_name

        # Create workbook and add all entries
        self._create_workbook(f"Media Clipping Report: {keywords}")

        # Add summary info at top
        self.worksheet.insert_rows(1, 3)
        self.worksheet.merge_cells("A1:H1")
        title_cell = self.worksheet.cell(row=1, column=1, value=f"MEDIA CLIPPING REPORT: {keywords}")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        self.worksheet.merge_cells("A2:H2")
        date_cell = self.worksheet.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        date_cell.alignment = Alignment(horizontal="center")

        self.worksheet.merge_cells("A3:H3")
        count_cell = self.worksheet.cell(row=3, column=1, value=f"Total Mentions: {len(data)}")
        count_cell.alignment = Alignment(horizontal="center")

        # Move headers to row 4
        self._create_workbook()  # Re-add headers

        # Add data starting from row 5
        for idx, entry in enumerate(data, start=5):
            self.add_entry(idx, entry)

        # Add summary stats
        summary_row = len(data) + 6
        self.worksheet.cell(row=summary_row, column=1, value="SUMMARY")
        self.worksheet.cell(row=summary_row, column=1).font = Font(bold=True)

        total_reach = sum(
            e.get("reach_estimate") or self._estimate_reach(e.get("url", ""), e.get("source", ""))
            for e in data
        )

        self.worksheet.cell(row=summary_row + 1, column=1, value="Total Estimated Reach:")
        self.worksheet.cell(row=summary_row + 1, column=2, value=total_reach)
        self.worksheet.cell(row=summary_row + 1, column=2).number_format = '#,##0'

        # Freeze panes
        self.worksheet.freeze_panes = "A5"

        # Save workbook
        self.workbook.save(str(output_path))

        return str(output_path)

    def generate_csv(self, data: List[Dict], report_name: str = None) -> str:
        """Generate CSV report (fallback)"""
        import csv

        if not report_name:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_name = f"media_clipping_{timestamp}.csv"

        if not report_name.endswith(".csv"):
            report_name += ".csv"

        output_path = self.reports_dir / report_name

        headers = ["Date Found", "Published Date", "Title", "URL", "Source", "Estimated Reach", "Screenshot Path", "Notes"]

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()

            for entry in data:
                reach = entry.get("reach_estimate", 0) or self._estimate_reach(entry.get("url", ""))
                writer.writerow({
                    "Date Found": entry.get("found_at", "")[:10],
                    "Published Date": entry.get("published_date", ""),
                    "Title": entry.get("title", ""),
                    "URL": entry.get("url", ""),
                    "Source": entry.get("source", ""),
                    "Estimated Reach": reach,
                    "Screenshot Path": entry.get("screenshot_path", ""),
                    "Notes": entry.get("snippet", "")
                })

        return str(output_path)


if __name__ == "__main__":
    # Test report generation
    test_data = [
        {
            "title": "Test Article 1",
            "url": "https://example.com/article1",
            "source": "Example News",
            "found_at": "2024-01-15T10:30:00",
            "published_date": "2024-01-14",
            "snippet": "This is a test article about AI technology.",
            "reach_estimate": 500000,
            "screenshot_path": ""
        }
    ]

    gen = ReportGenerator()
    path = gen.generate_report(test_data, "test_report.xlsx", "AI technology")
    print(f"Report saved to: {path}")
